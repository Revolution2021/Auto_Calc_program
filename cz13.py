import openpyxl as px
from datetime import datetime 
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl.styles import Font
from time import sleep

def totalcalc():

	wb1= px.load_workbook("modified1x_summary.xlsx",data_only=True)
	ws1=wb1.worksheets[0]
	ws1.cell(row=1,column=3).value="Units"
	ws1.cell(row=1,column=5).value="Unit Price"
	ws1.cell(row=1,column=6).value="Total Price(=Usage x Unit Price)"

	ws1max=ws1.max_row-1
	print(ws1max)

	for i1 in range(ws1max):
		val7a=ws1.cell(row=i1+2,column=3).value
		val7b=ws1.cell(row=i1+2,column=5).value
		print(val7a)
		val7x=val7a*val7b
		ws1.cell(row=i1+2,column=6).value=val7x
		ws1.cell(row=i1+2,column=6).number_format="#,##0"
		i1=i1+1

	wb1.save("modified2x_summary.xlsx")