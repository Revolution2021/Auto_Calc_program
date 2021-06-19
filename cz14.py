import openpyxl as px
from datetime import datetime 
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl.styles import Font
from time import sleep
from openpyxl.styles.alignment import Alignment


def decox():

	wb1= px.load_workbook("modified2x_summary.xlsx",data_only=True)
	ws1=wb1.worksheets[0]
	ws1max=ws1.max_row
	print(ws1max)
	FX="F"+str(ws1max)

	font1 = Font(color='00FF0000', size=12, italic=False, bold=True)
	font2 = Font(color='000000FF', size=12, italic=False, bold=False)
	font3 = Font(color='00FF0000', size=10, italic=False, bold=True)
	rightx = Alignment(horizontal='right', vertical='center')

	for i in range(ws1max-1):
		Fx='F'+str(i+2)
		print(Fx)
		ws1[Fx].font=font1
		ws1.cell(row=i+2,column=6).number_format="#,##0"
		s7=ws1.cell(row=i+2,column=6).value
#		ws1[Fx].value="USD  "+str(sz)
#		ws1.cell(row=i+2,column=6).value=f'{"USD"} {s7:,}'
		ws1.cell(row=i+2,column=6).value=f'{s7:,}'
		i=i+1

	i2=0
	while i2 < ws1max:	
		ws1.cell(row=i2+1,column=6).alignment = rightx
		i2=i2+1

	ws1['C1'].font=font2
	ws1['E1'].font=font2
	ws1['F1'].font=font2
	ws1.column_dimensions['D'].width=5
	ws1.column_dimensions['C'].width=11
	ws1.column_dimensions['E'].width=13
	ws1.column_dimensions['F'].width=32

	ws1.cell(row=1,column=3).alignment=rightx
	ws1.cell(row=1,column=5).alignment=rightx

	for f1x in glob.glob("ki*.csv"):
		os.remove(f1x)
			
	for f2 in glob.glob("final_data*.xlsx"):
		os.remove(f2)

	wb1.save("modified3x_summary.xlsx")