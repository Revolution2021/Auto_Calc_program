import openpyxl as px
from datetime import datetime 
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl.styles import Font
from time import sleep
import codecs

def totext():
	wb1= px.load_workbook("modified4x_summary.xlsx",data_only=True)
	ws1=wb1.worksheets[0]

	ws1max=ws1.max_row
	print(ws1max)

	font1 = Font(color='00FF0000', size=20, italic=True, bold=True)
	font2 = Font(color='000000FF', size=12, italic=False, bold=False)
	font3 = Font(color='00FF0000', size=10, italic=False, bold=True)

	ope='Tomas'
	pic='Jane'
	No=777
	pre_month=datetime.strftime(datetime.today()-relativedelta(months=1),"%Y/%m")

	print(f'Mr. {ope}、\n\nHi, This is {pic}.\n', file=codecs.open('samplex.txt', 'a', 'utf-8'))
	print(f'Billing charge to each client of previous month（{pre_month}）is as below.\n \n', file=codecs.open('samplex.txt', 'a', 'utf-8'))

	for i1 in range(ws1max):
		val1a=ws1.cell(row=i1+2,column=7).value  
#		val2a=ws1.cell(row=i1+2,column=8).value  
		val3a=ws1.cell(row=i1+2,column=2).value  
#	val3b=val3a[:11]   #kig*
		val4a=ws1.cell(row=i1+2,column=1).value #Company's name
#	val5x="val3b val4a"
		val6a=ws1.cell(row=i1+2,column=3).value  #Usage
		val7a=ws1.cell(row=i1+2,column=5).value # Unit Price
		val8a=ws1.cell(row=i1+2,column=6).value # Total Price
		val8s=f'{val8a} {"USD"}'
		print(val8s)
	
		if val1a is not None:
			text1=f'{val3a[:11]} : {val4a} unit number {val6a} x price/unit {val7a} USD ＝ Total_price {val8s} \n' 
#			text1=f'{val2a} \n{val3a[:11]} : {val4a} 数量 {val6a} x 単価 {val7a}円 ＝ 金額{val8a}円 \n' 
			print(text1, file=codecs.open('samplex.txt', 'a', 'utf-8'))
			print(text1)
	i1=i1+1

	print("Regards", file=codecs.open('samplex.txt', 'a', 'utf-8'))
	

	

