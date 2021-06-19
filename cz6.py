import openpyxl as px
from datetime import datetime 
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl.styles import Font

def finalsheet():
	filez6=glob.glob('final_*.xlsx')
	print(filez6)
	print(len(filez6))
	max6=len(filez6)
	print(max6)

	wb1= px.Workbook()
	ws1=wb1.worksheets[0]
	ws1.title="Final_Sheet"
		
	pre_month=datetime.strftime(datetime.today()-relativedelta(months=1),"%Y/%m")
	ws1.cell(1,1).value=pre_month+"__Billing_Charge" 
	ws1.column_dimensions['B'].width=45

	for i in range(max6):
		base, ext = os.path.splitext(filez6[i])

		wb2=px.load_workbook(filez6[i])
		print(base)
		ws2=wb2.active
		lis = [cell.value for cell in ws2["B:B"] if cell.value is not None]
		print(lis[-1])
	
		ws1.cell(i+2,2).value=base[22:]
		ws1.cell(i+2,3).value=lis[-1]
		ws1.cell(i+2,4).value="units"
	
		font1 = Font(color='00FF0000', size=20, italic=True, bold=True)
		ws1.cell(row=1,column=1).font = font1
		
		font2 = Font(color='000000FF', size=12, italic=True, bold=True)
		font3 = Font(color='00FF0000', size=10, italic=False, bold=True)
		cmax = ws2.max_row
		ws1.cell(row=i+2,column=2).font=font2
		ws1.cell(row=i+2,column=3).font=font3
		
		ws1.cell(row=i+2,column=3).number_format="#,##0"
		
	i=i+1

	
	ws1.column_dimensions['D'].width=5
	wb1.save("Summary.xlsx")

