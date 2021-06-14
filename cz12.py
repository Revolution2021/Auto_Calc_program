import openpyxl as px
from datetime import datetime 
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl.styles import Font
from time import sleep

def unitpriceset():
	wb1= px.load_workbook("modified_summary.xlsx",data_only=True)
	ws1=wb1.worksheets[0]
	wb2=px.load_workbook("client_list.xlsx",data_only=True)
	ws2=wb2["worksheet"]

	ws1max=ws1.max_row-1
	ws2max=ws2.max_row-1
	print(ws1max)
	print(ws2max)
	font1 = Font(color='00FF0000', size=20, italic=True, bold=True)
	font2 = Font(color='000000FF', size=12, italic=False, bold=False)
	font3 = Font(color='00FF0000', size=10, italic=False, bold=True)

	for i1 in range(ws1max):
		val3=ws1.cell(row=i1+2,column=2).value
		val3a=ws1.cell(row=i1+2,column=3).value
		print(val3)

		for i2  in range(ws2max):
			val4=ws2.cell(row=i2+2,column=3).value
			val4a=ws2.cell(row=i2+2,column=6).value
			val4b=ws2.cell(row=i2+2,column=7).value
			print(val3a)
			print(val4a)
			print(val4b)
			if val4==val3[:11] and val3a >=val4a and val4b >=val3a:
				ws1.cell(row=i1+2,column=5).value=ws2.cell(row=i2+2,column=9).value

		i2=i2+1
	i1=i1+1	

	ws1.column_dimensions['A'].width=45
	ws1.column_dimensions['A'].font=font2
	wb1.save("modified1x_summary.xlsx")