import openpyxl as px
from datetime import datetime 
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl.styles import Font
from time import sleep

def finalx():

	filez5=glob.glob('data_calc*.xlsx')
	print(filez5)
	print(len(filez5))
	max5=range(len(filez5))

	for i in max5:
		base, ext = os.path.splitext(filez5[i])
		wb1= px.load_workbook(base+"."+"xlsx",data_only=True)
		ws1=wb1.worksheets[0]
		print(ws1['J3'].value)
		wb1.save("RX"+base+"."+"xlsx")
		ws1=px.load_workbook("RX"+base+"."+"xlsx").active

		wb2=px.Workbook()
		sheet=wb2.active
		sheet.title="Summary"
		ws2=wb2["Summary"]
		ws2.column_dimensions['B'].width = 25

		pre_month=datetime.strftime(datetime.today()-relativedelta(months=1),"%Y/%m")
		print(pre_month)

		ws2.cell(row=1,column=1).value=pre_month
		ws2.cell(row=1,column=2).value=base[16:]+"summary"
		for i2 in range(1,10):

			copy = ws1.cell(row = i2, column = 10).value
			ws2.cell(row = i2, column = 2, value = copy)
			ws2.cell(row=i2,column=2).number_format = "#,##0.00"
			i2=i2+1
#			print(ws2.cell(row=4,column=2).value)
			wb2.save("final_"+base+"."+"xlsx")
			
			font1 = Font(color='00FF0000', size=25, italic=True, bold=True)
			ws2['B1'].font = font1
		
			font2 = Font(color='000000FF', size=25, italic=True, bold=True)
			cmax = ws2.max_row
			ws2.cell(row=cmax,column=2).font=font2

		i2=i2+1
	
	i=i+1



	for f1 in glob.glob("data_calc*.csv"):
		os.remove(f1)

	for f2 in glob.glob("copy*.xlsx"):
		os.remove(f2)

	for f3 in glob.glob("RX*.xlsx"):
		os.remove(f3)
	
	for f4 in glob.glob("copy*.csv"):
		os.remove(f4)
	
	for f5 in glob.glob("data_calc*.xlsx"):
		os.remove(f5)


