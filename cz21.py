import openpyxl as px
from datetime import datetime 
from dateutil.relativedelta import relativedelta
import os
import glob
from openpyxl.styles import Font
from time import sleep

def SGcodeadd():
	wb1= px.load_workbook("modified3x_summary.xlsx",data_only=True)
	ws1=wb1.worksheets[0]
 
	wb2=px.load_workbook("client_list.xlsx",data_only=True)
	ws2=wb2["worksheet"]

	ws1max=ws1.max_row-1
	print(ws1max)
	ws2max=ws2.max_row-1
	print(ws2max)

	font1 = Font(color='00FF0000', size=20, italic=True, bold=True)
	font2 = Font(color='000000FF', size=12, italic=False, bold=False)
	font3 = Font(color='00FF0000', size=10, italic=False, bold=True)

	for i1 in range(ws1max):
		val1=ws1.cell(row=i1+2,column=7).value  #Caution*
		val2=ws1.cell(row=i1+2,column=2).value  #kig*
		print(val2)
		val3=val2[:11]
		print(val3)
		val2x=ws1.cell(row=4,column=2).value
		val2y=val2x[:11]
		print(val2y)
		val12x=ws2.cell(row=20,column=3).value
		if val2y==val12x:
			print("BINGO!")
			
		val4=ws1.cell(row=i1+2,column=2) #Company's name
		val5=ws1.cell(row=i1+2,column=3)  #Usage
		val6=ws1.cell(row=i1+2,column=4) # Unit Price
		val7=ws1.cell(row=i1+2,column=5) # Total Price
	
		for i2 in range(ws2max):
			val10=ws2.cell(i2+2,column=3).value  #kig*
			val11=ws2.cell(i2+2,column=2).value  #SG No
			if val3==val10:
				ws1.cell(row=i1+2,column=8).value=val11
			
		i2=i2+1
		
	i1=i1+1	

	wb1.save("modified4x_summary.xlsx")

