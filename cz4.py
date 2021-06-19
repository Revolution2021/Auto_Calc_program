import openpyxl as px
import glob
import os
import math

def calcx():
	filez4=glob.glob('copy_XX*.xlsx')
	print(filez4)
	print(len(filez4))
	max4=range(len(filez4))

	for i in max4:
		base, ext = os.path.splitext(filez4[i])
		filepath = base+"."+"xlsx"
		wb = px.load_workbook(filename=filepath)
		ws1 = wb.worksheets[0]

# 最終行の取得
		cmax1 = ws1.max_row

# 計算用の変数goukeiを作成
		goukei_E = 0
		goukei_F = 0

# 所定範囲の数値を計算
		for i2 in range(2, cmax1+1):
			kingakuE = ws1['E' + str(i2)].value
			goukei_E += kingakuE

			kingakuF= ws1['F' + str(i2)].value
			goukei_F +=kingakuF
	
		ws1.cell(row=2,column=10).number_format = "#,##0.00"

		ws1.column_dimensions['J'].width = 25

# プログラム6｜合計値をセルJ2に出力
		ws1['J2'].value = goukei_E+goukei_F
		value1= "DUMMY_DATA"
		value2= "DUMMY_DATA"
		value3= ws1['J2'].value

		ws1.cell(row=3,column=10).value= value1
		ws1.cell(row=4,column=10).value= value2
		ws1.cell(row=5,column=10).value= value3


		ws1.title="calsh"
		wb.save("data_calc__"+base+"."+"xlsx")

	i=i+1